import math
import string
import warnings
import openpyxl

from datetime import datetime
from pathlib import Path
from collections import OrderedDict
from OEOrder import OEOrder

from accpac import getOrgPath

from extools.message import ExMessages
from extools import success
from extools.error_stack import consume_errors

__exm_eoi = None

LOG_FILE_NAME = "excel_order_importer.log"
DEBUG_FILE_NAME = "excel_order_importer.debug"

def format_errors(stack):
    return "\n- ".join([e[1].strip().replace('\n', ' ').replace('\r', '')
                        for e in stack])

def log():
    global __exm_eoi
    if not __exm_eoi:
        log_path = Path(getOrgPath(), LOG_FILE_NAME)
        debug_path = Path(getOrgPath(), DEBUG_FILE_NAME)
        debug = debug_path.exists()
        __exm_eoi = ExMessages("excel_order_importer",
                     level=ExMessages.DEBUG if debug else ExMessages.INFO,
                     log_path=str(log_path),
                     box=None
                     )
    return __exm_eoi

STRFTIME = "%Y%m%d"
STRPTIME = "%m/%d/%Y"

def parse_date(_in):
    try:
        return datetime.strptime(
                _in.split(" ")[0],
                STRPTIME).strftime(STRFTIME)
    except (ValueError, TypeError, IndexError, AttributeError) as e:
        return datetime.now().strftime(STRFTIME)

def parse_string(_in):
    if _in:
        try:
            return str(_in)
        except:
            pass
    return ""

HEADER = 'oeordh'
HEADER_OPTF = 'oeordho'
DETAIL = 'oeordd'
DETAIL_OPTF = 'oeorddo'

class ValidationError(RuntimeError):
    def __init__(self, field, value, message):
        self.field = field
        self.value = value
        self.message = message

    def __str__(self):
        return "Validation error for {} = {}: {}".format(
                self.field, self.value, self.message)

class ExcelOrderImporter():

    field_map = tuple()
    fixed_values = tuple()
    do_not_write = tuple()

    import_error_msg = "{ordnumber} {field} (Excel {row}, {col}): {error}"

    def __init__(self):
        self.orders_imported = []
        self.lines_imported = []
        self.errors = []
        self._min_required_fields()
        log().debug("Excel Order Imported initialized.")

    def _min_required_fields(self):
        defd_fields = [f[1] for f in self.field_map
                            if f[0] == HEADER]
        if "CUSTOMER" not in defd_fields:
            raise RuntimeError(
                    "CUSTOMER must be in the field_map.\n{}".format(
                    defd_fields))

    def open_worksheet(self, filepath):
        """Open and validate a worksheet.

        Opens the worksheet, sets it on the class, and runs
        ``worksheet_validators`` against it.

        :param path: a path-like object for the file to import.
        :type path: pathlib.Path
        :returns: (valid, [[obj, action, error_message], ... ])
        :rtype: (bool, [[str, str, str], ... ])
        """
        errors = []
        try:
            with warnings.catch_warnings():
                warnings.simplefilter("ignore")
                self.wb = openpyxl.load_workbook(
                        filepath.open('rb'), read_only=True, data_only=True)
        except Exception as err:
            msg = "Couldn't open workbook at {}: {}.".format(filepath, err)
            errors.append(['excel', 'open', msg, ])
            return (False, errors)

        if len(self.wb.worksheets) > 0:
            self.ws = self.wb.worksheets[0]
        else:
            msg = "Unable to find any worksheets in {}.".format(filepath)
            errors.append(['excel', 'open', msg, ])
            return (False, errors)

        self.ws.calculate_dimension(force=True)
        worksheet_validators = [method for method in self.__dir__()
                                if method.startswith('validate_worksheet')]

        for validator in worksheet_validators:
            if hasattr(self, validator):
                valid, msg = getattr(self, validator)(self.ws)
                if not valid:
                    errors.append(['excel', 'validate', msg])

        log().debug("Opened workbook {}. Errors: {}".format(filepath, errors))
        return (self.ws, errors, )

    def col_index_of(self, fieldname):
        for _v, field, index, _f in self.field_map:
            if field == fieldname:
                return index
        return -1

    def col_alpha_index(self, fieldname):
        index = self.col_index_of(fieldname)
        if index < 0:
            return "-1"
        aidx = ""
        divisor = index / len(string.ascii_uppercase)
        if divisor >= 1:
            didx = math.floor(divisor) - 1
            aidx += string.ascii_uppercase[didx]
        midx = index % len(string.ascii_uppercase)
        aidx += string.ascii_uppercase[midx]
        return aidx

    @property
    def header_writable_items(self):
        return [f[1] for f in self.field_map
                if f[0] == HEADER and f[1] not in self.do_not_write]
    @property
    def header_optf_writable_items(self):
        return [f[1] for f in self.field_map
                if f[0] == HEADER_OPTF and f[1] not in self.do_not_write]
    @property
    def detail_writable_items(self):
        return [f[1] for f in self.field_map
                if f[0] == DETAIL and f[1] not in self.do_not_write]
    @property
    def detail_optf_writable_items(self):
        return [f[1] for f in self.field_map
                if f[0] == DETAIL_OPTF and f[1] not in self.do_not_write]

    def clean(self, fields):
        return fields

    def validate_section(self, section, fields):
        for view, field, _, _ in self.field_map:
            if view == section:
                method_name = "validate_{}".format(field.lower())
                if hasattr(self, method_name):
                    valid, msg = getattr(self, method_name)(
                            fields[field], fields)
                    if not valid:
                        raise ValidationError(field, fields[field], msg)

    def import_xlsx(self, path):
        log().debug("Starting import from {}".format(path))
        ws, validation_errors = self.open_worksheet(path)
        if not ws:
            log().error("No worksheet after opening workbook: {}".format(
                validation_errors))
            return ([], [], validation_errors)

        errors = []
        orders_imported = []
        lines_imported = []
        errors = []

        # Iterate over the rows, skipping the header, and update the OEOrder.
        seen_header = False
        current_ordnumber = None

        # If a header cannot be imported, skip the lines.
        order_in_error = False

        order = OEOrder()

        for rowindex in range(1, self.ws.max_row + 1):
            row = ws[rowindex]

            if not seen_header:
                seen_header = True
                continue

            # Extract all the values from the row.
            fields = OrderedDict()
            for _v, field, index, func in self.field_map:
                try:
                    fields[field] = row[index].value
                except ValueError as e:
                    msg = self.import_error_msg.format(
                            ordnumber=fields.get('ORDNUMBER', "Unassigned"),
                            field=field,
                            row=rowindex, col=index, action="parse", error=e)
                    errors.append(("excel", "extract", msg, ))

            # Translate each value using the related type function
            for _v, field, index, func in self.field_map:
                try:
                    fields[field] = func(fields[field])
                except ValueError as e:
                    msg = self.import_error_msg.format(
                            row=rowindex, col=index,
                            ordnumber=fields.get('ORDNUMBER', "Unassigned"),
                            field=field,
                            action="translate", error=e)
                    errors.append(("excel", "translate", msg, ))

            orig_fields = fields.copy()

            # Use a cleaning pipeline on each field that needs sanitizing.
            for field in fields.keys():
                method_name = "clean_{}".format(field.lower())
                if hasattr(self, method_name):
                    msg = "Cleaning {}({}, {})".format(method_name, field, fields[field])
                    fields[field] = getattr(self, method_name)(
                            fields[field], fields)
                    log().debug(msg + " = {}".format(fields[field]))

            fields = self.clean(fields)

            stack = consume_errors()
            if stack:
                msg = self.import_error_msg.format(
                        row=rowindex, col=self.col_alpha_index("ORDNUMBER"),
                        ordnumber=fields['ORDNUMBER'],
                        field="",
                        action="clean",
                        error="{}".format(
                              format_errors(stack)))
                errors.append(("order", current_ordnumber, msg, ))

            if not fields.get("ORDNUMBER"):
                continue

            if fields['ORDNUMBER'] != current_ordnumber:
                current_ordnumber = fields['ORDNUMBER']
                order_in_error = False

                try:
                    self.validate_section(HEADER, fields)
                except ValidationError as e:
                    errors.append(('order', current_ordnumber, e.message))
                    order_in_error = True
                    continue

                # New order
                log().debug("New order for {}".format(fields["CUSTOMER"]))
                order = OEOrder()
                order.create(fields['CUSTOMER'])

                # Header fields
                for field in self.header_writable_items:
                    if fields.get(field) is not None:
                        order.oeordh.put(field, fields.get(field))

                for field, value in [(f[1], f[2], )
                         for f in self.fixed_values
                         if f[0] == HEADER and f[1] not in self.do_not_write]:
                    if value is not None and not success(
                            order.oeordh.put(field, value)):
                        stack = consume_errors()
                        msg = self.import_error_msg.format(
                            row=rowindex, col=self.col_alpha_index(field),
                            ordnumber=fields['ORDNUMBER'],
                            field=field,
                            action="put header value",
                            error="put {} = {}: {}".format(
                                field, value,
                                format_errors(stack)))
                        errors.append(("header",
                                       current_ordnumber,
                                       msg, ))

                try:
                    self.validate_section(HEADER_OPTF, fields)
                except ValidationError as e:
                    errors.append(
                            ('order optfield', current_ordnumber, e.message))

                # Header optional fields
                for field in self.header_optf_writable_items:
                    if fields.get(field) is None:
                        continue
                    order.oeordho.recordClear()
                    order.oeordho.put("OPTFIELD", field)
                    if success(order.oeordho.read()):
                        order.oeordho.put("VALUE", fields[field])
                        if not success(order.oeordho.update()):
                            stack = consume_errors()
                            msg = self.import_error_msg.format(
                                row=rowindex, col=self.col_alpha_index(field),
                                ordnumber=fields['ORDNUMBER'],
                                field=field,
                                action="update optional field",
                                error="update {}={}: {}".format(
                                    field, fields[field], format_errors(stack)))
                            errors.append(("optional field",
                                           current_ordnumber,
                                           msg, ))
                    else:
                        order.oeordho.recordGenerate()
                        order.oeordho.put("OPTFIELD", field)
                        order.oeordho.put("VALUE", fields[field])
                        if not success(order.oeordho.insert()):
                            msg = self.import_error_msg.format(
                                row=rowindex, col=self.col_alpha_index(field),
                                ordnumber=fields['ORDNUMBER'],
                                field=field,
                                action="insert optional field",
                                error="insert of {}={}".format(
                                        field, fields[field]))
                            errors.append(("optional field",
                                           current_ordnumber,
                                           msg, ))

                # Insert the order header
                ins = order.oeordh.insert()
                if success(ins):
                    orders_imported.append((orig_fields['ORDNUMBER'],
                                            order.oeordh.get("ORDNUMBER")))
                else:
                    order_in_error = True
                    stack = consume_errors()
                    msg = self.import_error_msg.format(
                            row=rowindex, col=self.col_alpha_index("ORDNUMBER"),
                            ordnumber=fields['ORDNUMBER'],
                            field="ORDNUMBER",
                            action="insert",
                            error="{}".format(
                                  format_errors(stack)))
                    errors.append(("order", current_ordnumber, msg, ))

            stack = consume_errors()
            if stack:
                msg = self.import_error_msg.format(
                        row=rowindex, col=self.col_alpha_index("ORDNUMBER"),
                        ordnumber=fields['ORDNUMBER'],
                        field="ORDNUMBER",
                        action="insert",
                        error="{}".format(
                              format_errors(stack)))
                errors.append(("order", current_ordnumber, msg, ))

            # if the order header import failed for this ID, skip all lines
            if order_in_error:
                continue

            # Handle Item
            try:
                self.validate_section(DETAIL, fields)
            except ValidationError as e:
                errors.append(
                        ('line', rowindex, e.message))
                continue

            rg = order.oeordd.recordGenerate()
            pi = order.oeordd.put("ITEM", fields['ITEM'])
            pip = order.oeordd.process()

            if not success(pi, pip):
                stack = consume_errors()
                msg = self.import_error_msg.format(
                            row=rowindex, col=self.col_alpha_index("ITEM"),
                            ordnumber=fields['ORDNUMBER'],
                            field=field,
                            action="insert error",
                            error="set item: {}".format(
                                  format_errors(stack)))
                errors.append(("line", current_ordnumber, msg, ))
                continue

            # Setup optional fields
            order.oeordd.put("PROCESSCMD", 1)
            order.oeordd.process()

            for field in self.detail_writable_items:
                if fields.get(field) is not None and not success(
                        order.oeordd.put(field, fields[field])):
                    stack = consume_errors()
                    msg = self.import_error_msg.format(
                        row=rowindex, col=self.col_alpha_index(field),
                        ordnumber=fields['ORDNUMBER'],
                        field=field,
                        action="put line value",
                        error="put {} = {}: {}".format(
                            field, value,
                            format_errors(stack)))
                    errors.append(("line",
                                   current_ordnumber,
                                   msg, ))

            for field, value in [(f[1], f[2], )
                                 for f in self.fixed_values
                                 if f[0] == DETAIL]:
                if value is not None:
                    order.oeordd.put(field, value)

            try:
                self.validate_section(DETAIL_OPTF, fields)
            except ValidationError as e:
                errors.append(
                        ('line', rowindex, e.message))

            for field in self.detail_optf_writable_items:
                if fields.get(field) is None:
                    continue
                order.oeorddo.recordClear()
                order.oeorddo.put("OPTFIELD", field)
                if success(order.oeorddo.read()):
                    order.oeorddo.put("VALUE", fields[field])
                    if not success(order.oeorddo.update()):
                        stack = consume_errors()
                        msg = self.import_error_msg.format(
                                row=rowindex, col=self.col_alpha_index(field),
                                ordnumber=fields['ORDNUMBER'],
                                field=field,
                                action="update",
                                error="Updating optfield {}: {}".format(
                                    field, format_errors(stack)))
                        errors.append(("line", rowindex, msg, ))
                else:
                    order.oeorddo.recordGenerate()
                    order.oeorddo.put("OPTFIELD", field)
                    order.oeorddo.put("VALUE", fields[field])
                    if not success(order.oeorddo.insert()):
                        stack = consume_errors()
                        msg = self.import_error_msg.format(
                                row=rowindex, col=self.col_alpha_index(field),
                                ordnumber=fields['ORDNUMBER'],
                                field=field,
                                action="insert",
                                error="Inserting optfield {}: {}".format(
                                    field, format_errors(stack)))
                        errors.append(("line", rowindex, msg, ))

            ins = order.oeordd.insert()

            if success(ins):
                lines_imported.append(
                        (current_ordnumber, fields.get('line', 0)))
                order.oeordh.update()
            else:
                stack = consume_errors()
                msg = self.import_error_msg.format(
                        row=rowindex, col=self.col_alpha_index("ORDNUMBER"),
                        ordnumber=fields['ORDNUMBER'],
                        field="ITEM",
                        action="insert",
                        error="Inserting line from row {}: {}.".format(
                            rowindex, format_errors(stack)))
                errors.append(("line", rowindex, msg, ))

            # Check if any warnings were raised
            stack = consume_errors()
            if stack:
                msg = self.import_error_msg.format(
                        row=rowindex, col=self.col_alpha_index("ORDNUMBER"),
                        ordnumber=fields['ORDNUMBER'],
                        field="ITEM",
                        action="insert",
                        error="Sage warnings: {}".format(format_errors(stack)))
                errors.append(("line", current_ordnumber, msg, ))

        self.wb.close()

        return orders_imported, lines_imported, errors
