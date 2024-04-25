#!/usr/bin/env python3
# encoding: utf-8
#
# This file is a part of the LinkAhead Project.
#
# Copyright (C) 2024 Indiscale GmbH <info@indiscale.com>
# Copyright (C) 2024 Henrik tom Wörden <h.tomwoerden@indiscale.com>
# Copyright (C) 2024 Daniel Hornung <d.hornung@indiscale.com>
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU Affero General Public License as
# published by the Free Software Foundation, either version 3 of the
# License, or (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU Affero General Public License for more details.
#
# You should have received a copy of the GNU Affero General Public License
# along with this program. If not, see <https://www.gnu.org/licenses/>.

from __future__ import annotations

import json
import pathlib
from collections import OrderedDict
from types import SimpleNamespace
from typing import Any, Dict, List, Optional, TextIO, Union
from warnings import warn

from jsonschema import FormatChecker, validate
from jsonschema.exceptions import ValidationError
from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.worksheet.worksheet import Worksheet

from .table_generator import ColumnType, RowType
from .utils import p2s


def _is_exploded_sheet(sheet: Worksheet) -> bool:
    """Return True if this is a an "exploded" sheet.

    An exploded sheet is a sheet whose data entries are LIST valued properties of entries in another
    sheet.  A sheet is detected as exploded iff it has FOREIGN columns.
    """
    column_types = _get_column_types(sheet)
    return ColumnType.FOREIGN.name in column_types.values()


def _get_column_types(sheet: Worksheet) -> OrderedDict:
    """Return an OrderedDict: column index -> column type for the sheet.
    """
    result = OrderedDict()
    type_row_index = _get_row_type_column_index(sheet)
    for idx, col in enumerate(sheet.columns):
        type_cell = col[type_row_index]
        result[idx] = type_cell.value if type_cell.value is not None else ColumnType.IGNORE.name
        assert (hasattr(ColumnType, result[idx])
                or result[idx] == RowType.COL_TYPE.name), (
            f"Unexpected column type value ({idx}{type_row_index}): {type_cell.value}")
    return result


def _get_foreign_key_columns(sheet: Worksheet) -> Dict[str, SimpleNamespace]:
    """Return the foreign keys of the worksheet.

Returns
-------
out: dict[str, SimpleNamespace]
  The keys are the stringified paths.  The values are SimpleNamespace objects with ``index``,
  ``path`` and ``column`` attributes.
    """
    column_types = _get_column_types(sheet)
    path_rows = _get_path_rows(sheet)
    result = OrderedDict()
    for for_idx, name in column_types.items():
        if name != ColumnType.FOREIGN.name:
            continue
        path = []
        for row in path_rows:
            component = sheet.cell(row=row+1, column=for_idx+1).value
            if component is None:
                break
            assert isinstance(component, str), f"Expected string: {component}"
            path.append(component)
        result[p2s(path)] = SimpleNamespace(index=for_idx, path=path,
                                            column=list(sheet.columns)[for_idx])
    return result


def _get_row_type_column_index(sheet: Worksheet):
    """Return the column index (0-indexed) of the column which defines the row types.
    """
    for col in sheet.columns:
        for cell in col:
            if cell.value == RowType.COL_TYPE.name:
                return cell.column - 1
    raise ValueError("The column which defines row types (COL_TYPE, PATH, ...) is missing")


def _get_path_rows(sheet: Worksheet):
    """Return the 0-based indices of the rows which represent paths."""
    rows = []
    rt_col = _get_row_type_column_index(sheet)
    for cell in list(sheet.columns)[rt_col]:
        if cell.value == RowType.PATH.name:
            rows.append(cell.row-1)
    return rows


def _next_row_index(sheet: Worksheet) -> int:
    """Return the index for the next data row.

    This is defined as the first row without any content.
    """
    return sheet.max_row


def _read_or_dict(data: Union[dict, str, TextIO]) -> dict:
    """If data is a json file name or input stream, read data from there."""
    if isinstance(data, dict):
        pass
    elif isinstance(data, str):
        with open(data, encoding="utf-8") as infile:
            data = json.load(infile)
    elif hasattr(data, "read"):
        data = json.load(data)
    else:
        raise ValueError(f"I don't know how to handle the datatype of `data`: {type(data)}")
    assert isinstance(data, dict)
    return data


class TemplateFiller:
    """Class to fill XLSX templates.  Has an index for all relevant columns."""

    def __init__(self, workbook: Workbook, graceful: bool = False):
        self._workbook = workbook
        self._graceful = graceful
        self._create_index()

    @property
    def workbook(self):
        return self._workbook

    def fill_data(self, data: dict):
        """Fill the data into the workbook."""
        self._handle_data(data=data)

    class Context:
        """Context for an entry: simple properties of all ancestors, organized in a dict.

        This is similar to a dictionary with all scalar element properties at the tree nodes up to
        the root.  Siblings in lists and dicts are ignored.  Additionally the context knows where
        its current position is.

        Lookup of elements can easily be achieved by giving the path (as ``list[str]`` or
        stringified path).

        """

        def __init__(self, current_path: List[str] = None, props: Dict[str, Any] = None):
            self._current_path = current_path if current_path is not None else []
            self._props = props if props is not None else {}  # this is flat

        def copy(self) -> TemplateFiller.Context:
            """Deep copy."""
            result = TemplateFiller.Context(current_path=self._current_path.copy(),
                                            props=self._props.copy())
            return result

        def next_level(self, next_level: str) -> TemplateFiller.Context:
            result = self.copy()
            result._current_path.append(next_level)  # pylint: disable=protected-access
            return result

        def __getitem__(self, path: Union[List[str], str], owner=None) -> Any:
            if isinstance(path, list):
                path = p2s(path)
            return self._props[path]

        def __setitem__(self, propname: str, value):
            fullpath = p2s(self._current_path + [propname])
            self._props[fullpath] = value

        def fill_from_data(self, data: Dict[str, Any]):
            # TODO recursive for dicts and list?
            """Fill current level with all scalar elements of ``data``."""
            for name, value in data.items():
                if not isinstance(value, (dict, list)):
                    self[name] = value
                elif isinstance(value, dict):
                    if not value or isinstance(list(value.items())[0], list):
                        continue
                    old_path = self._current_path
                    new_path = self._current_path.copy() + [name]
                    self._current_path = new_path
                    self.fill_from_data(data=value)
                    self._current_path = old_path

    def _create_index(self):
        """Create a sheet index for the workbook.

        Index the sheets by all path arrays leading to them.  Also create a simple column index by
        column type and path.

        """
        self._sheet_index = {}
        for sheetname in self._workbook.sheetnames:
            sheet = self._workbook[sheetname]
            type_column = [x.value for x in list(sheet.columns)[
                _get_row_type_column_index(sheet)]]
            # 0-indexed, as everything outside of sheet.cell(...):
            coltype_idx = type_column.index(RowType.COL_TYPE.name)
            path_indices = [i for i, typ in enumerate(type_column) if typ == RowType.PATH.name]

            # Get the paths, use without the leaf component for sheet indexing, with type prefix and
            # leaf for column indexing.
            for col_idx, col in enumerate(sheet.columns):
                if col[coltype_idx].value == RowType.COL_TYPE.name:
                    continue
                path = []
                for path_idx in path_indices:
                    if col[path_idx].value is not None:
                        path.append(col[path_idx].value)
                # col_key = p2s([col[coltype_idx].value] + path)
                # col_index[col_key] = SimpleNamespace(column=col, col_index=col_idx)
                if col[coltype_idx].value not in [ColumnType.SCALAR.name, ColumnType.LIST.name]:
                    continue

                path_str = p2s(path)
                assert path_str not in self._sheet_index
                self._sheet_index[path_str] = SimpleNamespace(
                    sheetname=sheetname, sheet=sheet, col_index=col_idx,
                    col_type=col[coltype_idx].value)

    def _handle_data(self, data: dict, current_path: List[str] = None,
                     context: TemplateFiller.Context = None,
                     only_collect_insertables: bool = False,
                     ) -> Optional[Dict[str, Any]]:
        """Handle the data and write it into ``workbook``.

Parameters
----------
data: dict
  The data at the current path position.  Elements may be dicts, lists or simple scalar values.

current_path: list[str], optional
  If this is None or empty, we are at the top level.  This means that all children shall be entered
  into their respective sheets and not into a sheet at this level.  ``current_path`` and ``context``
  must either both be given, or none of them.

context: TemplateFiller.Context, optional
  Directopry of scalar element properties at the tree nodes up to the root.  Siblings in lists
  and dicts are ignored.  ``context`` and ``current_path`` must either both be given, or none of
  them.

only_collect_insertables: bool, optional
  If True, do not insert anything on this level, but return a dict with entries to be inserted.


Returns
-------
out: union[dict, None]
  If ``only_collect_insertables`` is True, return a dict (path string -> value)
        """
        assert (current_path is None) is (context is None), (
            "`current_path` and `context` must either both be given, or none of them.")
        if current_path is None:
            current_path = []
        if context is None:
            context = TemplateFiller.Context()
        context.fill_from_data(data)

        insertables: Dict[str, Any] = {}
        for name, content in data.items():
            # TODO is this the best way to do it????
            if name == "file":
                continue
            path = current_path + [name]
            next_context = context.next_level(name)
            # preprocessing
            if isinstance(content, list):
                if not content:  # empty list
                    continue
                # List elements must be all of the same type.
                assert len(set(type(entry) for entry in content)) == 1

                if isinstance(content[0], dict):  # all elements are dicts
                    # An array of objects: must go into exploded sheet
                    for entry in content:
                        self._handle_data(data=entry, current_path=path, context=next_context)
                    continue
            elif isinstance(content, dict):  # we recurse and simply use the result
                if not current_path:  # Special handling for top level
                    self._handle_data(content, current_path=path, context=next_context)
                    continue
                insert = self._handle_data(content, current_path=path, context=next_context.copy(),
                                           only_collect_insertables=True)
                assert isinstance(insert, dict)
                assert not any(key in insertables for key in insert)
                insertables.update(insert)
                continue
            else:  # scalars
                content = [content]  # make list for unified treatment below

            # collecting the data
            assert isinstance(content, list)
            if len(content) > 1:
                content = [ILLEGAL_CHARACTERS_RE.sub("", str(x)) for x in content]
                value = ";".join(content)  # TODO we need escaping of values
            else:
                value = content[0]
                if isinstance(value, str):
                    value = ILLEGAL_CHARACTERS_RE.sub("", value)
            path_str = p2s(path)
            assert path_str not in insertables
            insertables[path_str] = value
        if only_collect_insertables:
            return insertables
        if not current_path:  # Top level returns, because there are only sheets for the children.
            return None

        # actual data insertion
        insert_row = None
        sheet = None
        for path_str, value in insertables.items():
            if self._graceful and path_str not in self._sheet_index:
                warn(f"Ignoring path with missing sheet index: {path_str}")
                continue
            sheet_meta = self._sheet_index[path_str]
            if sheet is None:
                sheet = sheet_meta.sheet
            assert sheet is sheet_meta.sheet, "All entries must be in the same sheet."
            col_index = sheet_meta.col_index
            if insert_row is None:
                insert_row = _next_row_index(sheet)

            sheet.cell(row=insert_row+1, column=col_index+1, value=value)

        # Insert foreign keys
        if insert_row is not None and sheet is not None and _is_exploded_sheet(sheet):
            try:
                foreigns = _get_foreign_key_columns(sheet)
            except ValueError:
                print(f"Sheet: {sheet}")
                raise
            for index, path in ((f.index, f.path) for f in foreigns.values()):
                value = context[path]
                sheet.cell(row=insert_row+1, column=index+1, value=value)

        return None


def fill_template(data: Union[dict, str, TextIO], template: str, result: str,
                  validation_schema: Union[dict, str, TextIO] = None) -> None:
    """Insert json data into an xlsx file, according to a template.

This function fills the json data into the template stored at ``template`` and stores the result as
``result``.

Parameters
----------
data: Union[dict, str, TextIO]
  The data, given as Python dict, path to a file or a file-like object.
template: str
  Path to the XLSX template.
result: str
  Path for the result XLSX.
validation_schema: dict, optional
  If given, validate the date against this schema first.  This raises an exception if the validation
  fails.  If no validation schema is given, try to ignore more errors in the data when filling the
  XLSX template.
"""
    data = _read_or_dict(data)
    assert isinstance(data, dict)

    # Validation
    if validation_schema is not None:
        validation_schema = _read_or_dict(validation_schema)
        try:
            validate(data, validation_schema, format_checker=FormatChecker())
        except ValidationError as ve:
            print(ve.message)
            raise ve
    else:
        print("No validation schema given, continue at your own risk.")

    # Filling the data
    result_wb = load_workbook(template)
    template_filler = TemplateFiller(result_wb, graceful=(validation_schema is None))
    template_filler.fill_data(data=data)

    parentpath = pathlib.Path(result).parent
    parentpath.mkdir(parents=True, exist_ok=True)
    result_wb.save(result)
