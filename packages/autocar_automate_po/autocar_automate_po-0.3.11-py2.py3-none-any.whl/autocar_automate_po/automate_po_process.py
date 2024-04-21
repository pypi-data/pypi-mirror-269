try:
    import openpyxl as pyxl
    from openpyxl.worksheet.datavalidation import DataValidation
    import pandas as pd
    # import xlwings as xw
    import pyperclip
except ImportError:
    print('Please install the following dependencies:')
    print('openpyxl')
    print('pandas')
    print('xlwings')
    print('pyperclip')
    print('\nUse the following command:')
    print('pip install -r requirements.txt --upgrade')
    _ = input('Press Enter to exit\n')
    exit()

# Standard
from pathlib import Path
from datetime import datetime
import re
import warnings
from os import getlogin
from shutil import move

# Test server
TEST_MODE = False

# Disable warnings
warnings.filterwarnings('ignore')

# Display full dataframe
pd.set_option('display.max_rows', None)
pd.set_option('display.max_columns', None)

# Dataframe structure
DF = pd.DataFrame(columns=[
    'PO_Identifier',  # A
    'File_Name',  # B
    'Supplier_From_PR',  # C
    'Supplier_Number',  # D
    'Supplier_Name',  # E
    'Ship_To_Text',  # F
    'Site',  # G
    'Ship_To',  # H
    'PR_Date',  # I
    'Date_Required',  # J
    'Quantity',  # K
    'UOM',  # L
    'Part_Number',  # M
    'Description',  # N
    'Unit_Price',  # O
    'Purchase_Account',  # P
    'Sub_Account',  # Q
    'Cost_Center',  # R
    'Short_Requested_By',  # S
    'Requested_By',  # T
    'Line_Comments',  # U
    'Header_Comments',  # V
    'Remarks'  # W
]
)

try:
    df_supplier = pd.read_excel(Path('input') / 'supplier_list.xlsx')
except Exception:
    try:
        df_supplier = pd.read_csv(Path('input') / 'supplier_list.csv')
    except Exception:
        print('Please ensure that the file "supplier_list" is in the input folder')
        exit()
df_supplier['Lower Name'] = df_supplier['Sort Name'].str.lower()

COLS = [
    'A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J',
    'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W'
]


def main():
    print("""
██████╗░░█████╗░  ░█████╗░██╗░░░██╗████████╗░█████╗░███╗░░░███╗░█████╗░████████╗██╗░█████╗░███╗░░██╗
██╔══██╗██╔══██╗  ██╔══██╗██║░░░██║╚══██╔══╝██╔══██╗████╗░████║██╔══██╗╚══██╔══╝██║██╔══██╗████╗░██║
██████╔╝██║░░██║  ███████║██║░░░██║░░░██║░░░██║░░██║██╔████╔██║███████║░░░██║░░░██║██║░░██║██╔██╗██║
██╔═══╝░██║░░██║  ██╔══██║██║░░░██║░░░██║░░░██║░░██║██║╚██╔╝██║██╔══██║░░░██║░░░██║██║░░██║██║╚████║
██║░░░░░╚█████╔╝  ██║░░██║╚██████╔╝░░░██║░░░╚█████╔╝██║░╚═╝░██║██║░░██║░░░██║░░░██║╚█████╔╝██║░╚███║
╚═╝░░░░░░╚════╝░  ╚═╝░░╚═╝░╚═════╝░░░░╚═╝░░░░╚════╝░╚═╝░░░░░╚═╝╚═╝░░╚═╝░░░╚═╝░░░╚═╝░╚════╝░╚═╝░░╚══╝
    """)
    # Declare dataframe
    df = DF
    files_not_opened = []
    # Get unique identifier
    time_identifier = datetime.now().strftime('%m%d%y-%H%M%S')
    i_identifier = 1
    files_to_process = 0
    for path in Path('Files_To_Process').glob('*.xlsx'):
        files_to_process += 1
    # Loop Excel 2003 files
    excel_2003_files = []
    for path in Path('Files_To_Process').glob('*.xls'):
        excel_2003_files.append(path.stem)
    # Loop po files
    counter = 1
    for path in Path('Files_To_Process').glob('*.xlsx'):
        print_progress_bar(counter, files_to_process)
        try:
            _ = pyxl.load_workbook(filename=path)
            df = df.append(
                load_file_data(path=path, po_identifier=time_identifier + '-' + str(i_identifier), file_name=path.stem)
            )
        except PermissionError:
            files_not_opened.append(path.stem)
        counter += 1
        i_identifier += 1
    # Check if df is empty
    if df.empty:
        print('Error: No files loaded')
        print('Check that Purchase Requisitions are in the correct format (xlsx)')
        _ = input('Press Enter to exit')
        exit()
    # Export
    Path('export').mkdir(exist_ok=True)
    check_filename = Path('export') / ('Check_Export-' + time_identifier + '.xlsx')
    df.to_excel(check_filename, index=False)

    # Load check sheet
    wb_check = pyxl.load_workbook(filename=check_filename)
    ws_check = wb_check.active
    lr_ws_check = len(ws_check['A'])

    # Get columns
    unit_price_col = ''
    supplier_nr_col = ''
    site_col = ''
    ship_to_col = ''
    pr_date_col = ''
    date_req_col = ''
    quantity_col = ''
    uom_col = ''
    part_nr_col = ''
    description_col = ''
    purchase_acnt_col = ''
    sub_acnt_col = ''
    cost_center_col = ''
    short_req_by_col = ''
    comments_col = ''
    supplier_pr_col = ''
    supplier_name_col = ''
    ship_to_text_col = ''
    requested_by_col = ''
    line_comments_col = ''
    remarks_col = ''

    for col in COLS:
        col_val = str(ws_check[col + '1'].value).lower()
        if col_val == 'unit_price':
            unit_price_col = col
        if col_val == 'supplier_number':
            supplier_nr_col = col
        if col_val == 'site':
            site_col = col
        if col_val == 'ship_to':
            ship_to_col = col
        if col_val == 'pr_date':
            pr_date_col = col
        if col_val == 'date_required':
            date_req_col = col
        if col_val == 'quantity':
            quantity_col = col
        if col_val == 'uom':
            uom_col = col
        if col_val == 'part_number':
            part_nr_col = col
        if col_val == 'description':
            description_col = col
        if col_val == 'unit_price':
            unit_price_col = col
        if col_val == 'purchase_account':
            purchase_acnt_col = col
        if col_val == 'sub_account':
            sub_acnt_col = col
        if col_val == 'cost_center':
            cost_center_col = col
        if col_val == 'short_requested_by':
            short_req_by_col = col
        if col_val == 'header_comments':
            comments_col = col
        if col_val == 'supplier_from_pr':
            supplier_pr_col = col
        if col_val == 'supplier_name':
            supplier_name_col = col
        if col_val == 'ship_to_text':
            ship_to_text_col = col
        if col_val == 'requested_by':
            requested_by_col = col
        if col_val == 'line_comments':
            line_comments_col = col
        if col_val == 'remarks':
            remarks_col = col

    # Conditional formatting
    yellow_fill = pyxl.styles.PatternFill(bgColor='00FFFF00')
    dxf_yellow = pyxl.styles.differential.DifferentialStyle(fill=yellow_fill)
    conditional_formatting_rule_yellow = pyxl.formatting.Rule(type="expression", dxf=dxf_yellow, stopIfTrue=False)
    red_fill = pyxl.styles.PatternFill(bgColor='00FF0000')
    dxf_red = pyxl.styles.differential.DifferentialStyle(fill=red_fill)
    conditional_formatting_rule_red1 = pyxl.formatting.Rule(type="expression", dxf=dxf_red, stopIfTrue=False)
    conditional_formatting_rule_red2 = pyxl.formatting.Rule(type="expression", dxf=dxf_red, stopIfTrue=False)
    conditional_formatting_rule_red3 = pyxl.formatting.Rule(type="expression", dxf=dxf_red, stopIfTrue=False)
    conditional_formatting_rule_yellow.formula = ['C2=""']  # HARDCODED
    conditional_formatting_rule_red1.formula = ['K2=0']  # HARDCODED
    conditional_formatting_rule_red2.formula = ['O2=0']  # HARDCODED
    conditional_formatting_rule_red3.formula = ['=LEN(M2)>18']  # HARDCODED
    ws_check.conditional_formatting.add('C2:' + COLS[-4] + str(lr_ws_check), conditional_formatting_rule_yellow)  # HARDCODED
    ws_check.conditional_formatting.add(quantity_col + '2:' + quantity_col + str(lr_ws_check), conditional_formatting_rule_red1)
    ws_check.conditional_formatting.add(unit_price_col + '2:' + unit_price_col + str(lr_ws_check), conditional_formatting_rule_red2)
    ws_check.conditional_formatting.add(part_nr_col + '2:' + part_nr_col + str(lr_ws_check), conditional_formatting_rule_red3)

    # Set Number format
    for row in range(1, lr_ws_check + 1):
        ws_check[unit_price_col + str(row)].number_format = '$ #,##0.0000'
        ws_check[supplier_nr_col + str(row)].number_format = '@'
        ws_check[part_nr_col + str(row)].number_format = '@'

    # Highlight columns
    grey_color = pyxl.styles.colors.Color(rgb='00C0C0C0')
    grey_fill = pyxl.styles.fills.PatternFill(patternType='solid', fgColor=grey_color)
    edit_color = pyxl.styles.colors.Color(rgb='0099CCFF')
    edit_fill = pyxl.styles.fills.PatternFill(patternType='solid', fgColor=edit_color)

    # Set colors
    for row in range(1, lr_ws_check + 1):
        ws_check['A' + str(row)].fill = grey_fill
        ws_check['B' + str(row)].fill = grey_fill
        ws_check[supplier_pr_col + str(row)].fill = grey_fill
        ws_check[supplier_name_col + str(row)].fill = grey_fill
        ws_check[ship_to_text_col + str(row)].fill = grey_fill
        ws_check[requested_by_col + str(row)].fill = edit_fill
        ws_check[line_comments_col + str(row)].fill = edit_fill
        ws_check[remarks_col + str(row)].fill = grey_fill

        ws_check[supplier_nr_col + str(row)].fill = edit_fill
        ws_check[site_col + str(row)].fill = edit_fill
        ws_check[ship_to_col + str(row)].fill = edit_fill
        ws_check[pr_date_col + str(row)].fill = edit_fill
        ws_check[date_req_col + str(row)].fill = edit_fill
        ws_check[quantity_col + str(row)].fill = edit_fill
        ws_check[uom_col + str(row)].fill = edit_fill
        ws_check[part_nr_col + str(row)].fill = edit_fill
        ws_check[description_col + str(row)].fill = grey_fill
        ws_check[unit_price_col + str(row)].fill = edit_fill
        ws_check[purchase_acnt_col + str(row)].fill = edit_fill
        ws_check[sub_acnt_col + str(row)].fill = edit_fill
        ws_check[cost_center_col + str(row)].fill = edit_fill
        ws_check[short_req_by_col + str(row)].fill = edit_fill
        ws_check[comments_col + str(row)].fill = grey_fill

    # Borders under new PO
    border = pyxl.styles.borders.Border(bottom=pyxl.styles.borders.Side(style='thick'))
    for row in range(1, lr_ws_check + 1):
        if ws_check['A' + str(row)].value != ws_check['A' + str(row + 1)].value:
            for col in COLS:
                ws_check[col + str(row)].border = border

    # Data validation
    # UOM
    dv_uom = DataValidation(type='list', formula1='"EA,FT,QT,PT,PC,OZ,MM,ML,LB,in,GL,GA,RL,YD"', allow_blank=True)
    ws_check.add_data_validation(dv_uom)
    dv_uom.add(uom_col + '2:' + uom_col + str(lr_ws_check))

    # Freeze panes
    ws_check.freeze_panes = ws_check['A2']

    # Save check file
    wb_check.save(filename=check_filename)

    # Xlwings autofit
    # xw_wb = xw.Book(check_filename)
    # xw_ws = xw_wb.sheets.active
    # xw_ws.autofit(axis='columns')
    # xw_wb.save(check_filename)

    # Files not opened
    if files_not_opened:
        print('Files not opened:')
        print(*files_not_opened, sep='\n')
    # Excel 2003 files
    if excel_2003_files:
        print('Excel 2003 (xls) files detected - Please convert these files:')
        print(*excel_2003_files, sep='\n')

    # Ask user to check and update the file
    user_input_check_file('\nPlease check and save the file: ' + check_filename.stem)

    # Populate df_export
    populate_export(check_filename, time_identifier)

    # Ask if import was successful
    while True:
        in_import_suc = input('\nWas the import successful? Press "Y" to continue and "N" after the errors are fixed in the file: ' + check_filename.stem + '\n')
        if in_import_suc.lower() == 'y':
            break
        if in_import_suc.lower() == 'n':
            populate_export(check_filename, time_identifier)
        else:
            print('Invalid entry')
            continue

    # Copy the PO file to server
    po_filename = ('pr_input_' + getlogin() + '_' + time_identifier + '_PR_output' + '.csv')
    po_new_filename = po_filename[-20:]
    while True:
        user_input_check_file('\nCopy the Requisition Import Success file: ' + po_filename + ' to this folder')
        try:
            if TEST_MODE:
                move(po_filename, Path('//gvwac52/users/requisition/PO/import') / po_new_filename)
            else:
                move(po_filename, Path('//gvwac53/users/requisition/PO/import') / po_new_filename)
            break
        except Exception:
            print('\nUnable to move the file. Please ensure that the file: ' + po_filename + ' is copied to this folder')
            continue
    pyperclip.copy(po_new_filename)
    print('\n--------------------PO IMPORT--------------------')
    print('Copied to clipboard: ' + po_new_filename)
    # ----------------------------------------------------------------------------------------------------
    _ = input('Press Enter to exit\n')


def load_file_data(path, po_identifier, file_name):
    wb = pyxl.load_workbook(filename=path, data_only=True)
    ws = wb.active
    # Get last row
    last_row_a = len(ws['A'])
    last_row_g = len(ws['G'])
    last_row = max(last_row_a, last_row_g) + 10

    # --------------------Header data--------------------
    # Date
    try:
        if ws['G4'].value is not None:
            try:
                head_date = ws['G4'].value
                if head_date < datetime.today():
                    head_date = datetime.today()
                    head_date = head_date.strftime('%m/%d/%Y')
                else:
                    head_date = head_date.strftime('%m/%d/%Y')
            except AttributeError:
                head_date = datetime.today()
                head_date = head_date.strftime('%m/%d/%Y')
        elif ws['H4'].value is not None:
            try:
                head_date = ws['H4'].value
                if head_date < datetime.today():
                    head_date = datetime.today()
                    head_date = head_date.strftime('%m/%d/%Y')
                else:
                    head_date = head_date.strftime('%m/%d/%Y')
            except AttributeError:
                head_date = datetime.today()
                head_date = head_date.strftime('%m/%d/%Y')
        else:
            head_date = ''
    except Exception:
        head_date = ''

    # Supplier name from PR
    head_supplier_name = ''
    try:
        if ws['A8'].value is not None:
            head_supplier_name = ws['A8'].value
        elif ws['A9'].value is not None:
            head_supplier_name = ws['A9'].value
        elif ws['A14'].value is not None:
            head_supplier_name = ws['A14'].value
        elif ws['A15'].value is not None:
            head_supplier_name = ws['A15'].value
    except Exception:
        head_supplier_name = ''
    # Supplier number and name guess
    try:
        df_found = df_supplier[df_supplier['Lower Name'].str.contains(str(head_supplier_name).lower())]
        supplier_nr = str(df_found.iloc[0, 0])
        supplier_name = df_found.iloc[0, 1]
    except IndexError:
        supplier_nr = ''
        supplier_name = ''
    # Ship to address
    try:
        ship_to_address = ''
        ship_to_row = 0
        ship_to_data_row = 0
        for row in range(1, last_row):
            if str(ws['E' + str(row)].value).lower() == 'ship to address':
                ship_to_row = row + 1
                break
        for row in range(ship_to_row, last_row):
            if ws['E' + str(row)].value is not None:
                ship_to_data_row = row
                break
        for row in range(ship_to_data_row, last_row):
            if ws['E' + str(row)].value is not None:
                ship_to_address += ws['E' + str(row)].value + ' '
            else:
                break
    except Exception:
        ship_to_address = ''
    # Site
    site = ''
    ship_to_nr = ''
    if re.search('pinson|birmingham', ship_to_address.lower()):
        site = '4000'
        ship_to_nr = '10005773'
    if re.search('washington|hagerstown', ship_to_address.lower()):
        site = '2000'
        ship_to_nr = '20000000'
    # Remarks
    try:
        remarks_row = 0
        for row in range(20, last_row):
            if str(ws['A' + str(row)].value).lower() == 'remarks/special packaging instructions:':
                remarks_row = row + 1
                break
        remarks = ws['A' + str(remarks_row)].value
    except Exception:
        remarks = ''
    # Requested by
    try:
        requested = ''
        requested_row = 0
        for row in range(20, last_row):
            if str(ws['A' + str(row)].value).lower() == 'requested by:':
                requested_row = row
                break
        if ws['A' + str(requested_row + 1)].value is not None:
            requested = ws['A' + str(requested_row + 1)].value
        elif ws['A' + str(requested_row + 2)].value is not None:
            requested = ws['A' + str(requested_row + 2)].value
        elif ws['B' + str(requested_row)].value is not None:
            requested = ws['B' + str(requested_row)].value
    except Exception:
        requested = ''
    # Short requested by
    short_requested = ''
    _short_requested = ''
    if requested != '':
        try:
            _short_requested = str(requested[:1]).upper() + str(requested).split(' ')[1].upper()
            _short_requested = _short_requested[0:8]
            for char in _short_requested:
                if char.isalnum():
                    short_requested += char
        except IndexError:
            _short_requested = str(requested).upper()[0:8]
            for char in _short_requested:
                if char.isalnum():
                    short_requested += char
    # --------------------Line data--------------------
    data_start_row = 0
    for row in range(1, last_row):
        if str(ws['B' + str(row)].value).lower() == 'date required':
            data_start_row = row + 1
            break
    if data_start_row == 0:
        for row in range(1, last_row):
            if str(ws['C' + str(row)].value).lower() == 'date required':
                data_start_row = row + 1
                break
    if data_start_row == 0:
        for row in range(1, last_row):
            if str(ws['C' + str(row)].value).lower() == 'quantity':
                data_start_row = row + 1
                break
    if data_start_row == 0:
        for row in range(1, last_row):
            if str(ws['D' + str(row)].value).lower() == 'quantity':
                data_start_row = row + 1
                break

    data_end_row = 0
    for row in range(data_start_row, last_row):
        if ws['B' + str(row)].value is None:
            data_end_row = row - 1
            break
    # Check if Column B (Date Required) is empty
    if data_end_row < data_start_row:
        for row in range(data_start_row, last_row):
            if ws['C' + str(row)].value is None:
                data_end_row = row - 1
                break
    # TODO: add check for data_start_row and data_end_row = 0
    # Get columns
    for i in range(len(COLS)):
        i_val = str(ws[COLS[i] + str(data_start_row - 1)].value).lower()
        if i_val == 'date required':
            col_date_required = COLS[i]
        if i_val == 'quantity':
            col_quantity = COLS[i]
        if i_val == 'part number':
            col_part_number = COLS[i]
        if i_val == 'description':
            col_description = COLS[i]
        if i_val == 'unit price':
            col_unit_price = COLS[i]
    # Add data to dataframe
    df = DF
    for row in range(data_start_row, data_end_row + 1):
        try:
            date_required = ws[col_date_required + str(row)].value
            if date_required < datetime.today():
                date_required = datetime.today()
                date_required = date_required.strftime('%m/%d/%Y')
            else:
                date_required = date_required.strftime('%m/%d/%Y')
        except Exception:
            date_required = datetime.today()
            date_required = date_required.strftime('%m/%d/%Y')
        try:
            quantity = ws[col_quantity + str(row)].value
        except Exception:
            quantity = ''
        try:
            part_number = str(ws[col_part_number + str(row)].value)
            part_number = part_number.strip()
            part_number = part_number.replace('"', '')
        except Exception:
            part_number = ''
        try:
            description = str(ws[col_description + str(row)].value)
        except Exception:
            description = ''
        try:
            unit_price = ws[col_unit_price + str(row)].value
            unit_price = round(unit_price, 5)
        except Exception:
            unit_price = ''

        line_comments = ''
        try:
            if remarks is not None:
                if remarks[0:2] == 'SD':
                    line_comments = remarks
        except Exception:
            remarks = ''

        data_to_df = {
            'PO_Identifier': po_identifier,
            'File_Name': file_name,
            'Date_Required': date_required,
            'Quantity': quantity,
            'Part_Number': part_number,
            'Description': description,
            'Unit_Price': unit_price,
            'Remarks': remarks,
            'PR_Date': head_date,
            'Supplier_From_PR': head_supplier_name,
            'Requested_By': requested,
            'Ship_To_Text': ship_to_address,
            'Site': site,
            'UOM': 'EA',
            'Short_Requested_By': short_requested,
            'Purchase_Account': '',
            'Sub_Account': '',
            'Cost_Center': '',
            'Line_Comments': line_comments,
            'Ship_To': ship_to_nr,
            'Header_Comments': 'Please reference this Purchase Order on the Invoice. Email the Invoice to ap@autocartruck.com',
            'Supplier_Number': supplier_nr,
            'Supplier_Name': supplier_name
        }
        df = df.append(data_to_df, ignore_index=True)
    return df


def populate_export(check_filename, time_identifier):
    def df_check_read_excel():
        df_check = pd.read_excel(check_filename, converters={
            'Supplier_Number': str,
            'Part_Number': str,
            'Site': str,
            'Ship_To': str,
            'Purchase_Account': str,
            'Sub_Account': str,
            'Cost_Center': str,
            'Requested_By': str,
            'Line_Comments': str,
            'Header_Comments': str
        })
        return df_check
    df_check = df_check_read_excel()
    # Change date formats
    while True:
        try:
            df_check['PR_Date'] = pd.to_datetime(df_check['PR_Date'])
        except Exception:
            user_input_check_file('\nPlease check the PR_Date column and save the file when complete')
            df_check = df_check_read_excel()
        else:
            break
    df_check['PR_Date'] = df_check['PR_Date'].dt.strftime('%m/%d/%y')

    while True:
        try:
            df_check['Date_Required'] = pd.to_datetime(df_check['Date_Required'])
        except Exception:
            user_input_check_file('\nPlease check the Date_Required column and save the file when complete')
            df_check = df_check_read_excel()
        else:
            break
    df_check['Date_Required'] = df_check['Date_Required'].dt.strftime('%m/%d/%y')

    # Set requested by comment
    df_check['Comment_Requested'] = 'Requested by ' + df_check['Requested_By'].astype(str) + '. '

    # Part_Number length 18 char
    df_check['Part_Number'] = df_check['Part_Number'].str.slice(0, 18)

    # Populate export
    df_export = pd.DataFrame()
    df_export['Request ID'] = df_check['PO_Identifier']
    df_export['Item'] = df_check['Part_Number'].str.strip()
    df_export['Description'] = df_check['Description']
    df_export['Site'] = df_check['Site'].str.strip()
    df_export['Quantity'] = df_check['Quantity']
    df_export['UM'] = df_check['UOM']
    df_export['Unit Cost'] = df_check['Unit_Price']
    df_export['Release Date'] = df_check['PR_Date']
    df_export['Need Date'] = df_check['Date_Required']
    df_export['Requested By'] = df_check['Short_Requested_By'].str.strip()
    df_export['Pur Acct'] = df_check['Purchase_Account'].str.strip()
    df_export['Sub Acct'] = df_check['Sub_Account'].str.strip()
    df_export['Cost Center'] = df_check['Cost_Center'].str.strip()
    df_export['Supplier'] = df_check['Supplier_Number'].str.strip()
    df_export['Ship-To'] = df_check['Ship_To'].str.strip()
    df_export['Header Comments'] = df_check['Comment_Requested'] + df_check['Header_Comments']
    df_export['Line Comments'] = df_check['Line_Comments']

    print('--------------------PR Import Data--------------------')
    df_print = df_export[['Item', 'Site', 'Quantity', 'UM', 'Unit Cost', 'Need Date', 'Supplier']]
    print(df_print)
    print('--------------------PR Import Data--------------------')
    user_input_check_file('\nAre you sure you want to import?')
    export_filename = ('pr_input_' + getlogin() + '_' + time_identifier + '.csv')
    if TEST_MODE:
        df_export.to_csv(Path('//gvwac52/users/requisition/PR/import') / export_filename, index=False)
    else:
        df_export.to_csv(Path('//gvwac53/users/requisition/PR/import') / export_filename, index=False)
    pyperclip.copy(export_filename)
    print('\n--------------------PO REQUISITION--------------------')
    print('Copied to clipboard: ' + export_filename)


def print_progress_bar(iteration, total, prefix='', suffix='', decimals=0, length=100, fill='█', printEnd='\r'):
    """
    Call in a loop to create terminal progress bar
    @params:
        iteration   - Required  : current iteration (Int)
        total       - Required  : total iterations (Int)
        prefix      - Optional  : prefix string (Str)
        suffix      - Optional  : suffix string (Str)
        decimals    - Optional  : positive number of decimals in percent complete (Int)
        length      - Optional  : character length of bar (Int)
        fill        - Optional  : bar fill character (Str)
        printEnd    - Optional  : end character (e.g. "\r", "\r\n") (Str)
    """
    percent = ("{0:." + str(decimals) + "f}").format(100 * (iteration / float(total)))
    filledLength = int(length * iteration // total)
    bar = fill * filledLength + '-' * (length - filledLength)
    print(f'\r{prefix} |{bar}| {percent}% {suffix}', end=printEnd)
    # Print New Line on Complete
    if iteration == total:
        print()


def user_input_check_file(prompt):
    print(prompt)
    input_checked = input('Press "Y" when complete and "N" to exit\n')
    while True:
        if input_checked.lower() == 'y':
            break
        elif input_checked.lower() == 'n':
            _ = input('Press Enter to exit\n')
            exit()
        else:
            print('Invalid entry')
            input_checked = input('Press "Y" when complete and "N" to exit\n')
            continue


if __name__ == '__main__':
    main()
