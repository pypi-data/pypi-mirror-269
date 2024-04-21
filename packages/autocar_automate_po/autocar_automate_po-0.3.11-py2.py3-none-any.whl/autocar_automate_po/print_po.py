from pathlib import Path
from datetime import datetime
import os
from time import sleep
from shutil import copy, move

try:
    import openpyxl as pyxl
    from openpyxl.styles import Alignment
except ImportError:
    print('Please install the following dependencies:')
    print('openpyxl')
    _ = input('Press Enter to exit\n')
    exit()

# Network location
NETWORK_SAVE = r'\\gvwac09\Public\Finance\Purchase Orders\__To Process__'


def main():
    # Declare variables
    files_not_opened = []
    files_to_move = []
    files_not_moved = []
    # Get unique identifier
    time_identifier = datetime.now().strftime('%m%d%y-%H%M%S')
    i_identifier = 1

    Path('export').mkdir(exist_ok=True)
    Path('complete').mkdir(exist_ok=True)

    for path in Path('Files_To_Process').glob('*.xlsx'):
        try:
            po_identifier = 'PR-' + time_identifier + '-' + str(i_identifier)
            change_file(path, po_identifier)
            print('Printing file: ' + po_identifier)
            os.startfile(Path('export') / (po_identifier + '.xlsx'), 'print')
            files_to_move.append(po_identifier + '.xlsx')
            print('Loading...')
            sleep(10)
            i_identifier += 1
        except Exception:
            files_not_opened.append(path.stem)

    print('Moving files to P drive...')
    for path in files_to_move:
        try:
            copy(Path('export') / path, Path(NETWORK_SAVE) / path)
        except Exception:
            files_not_moved.append(path)

    for path in Path('Files_To_Process').glob('*.xlsx'):
        try:
            move(path, Path('complete') / path.name)
        except Exception:
            files_not_moved.append(path)

    if files_not_opened:
        print('Files not opened:')
        print(*files_not_opened, sep='\n')

    if files_not_moved:
        print('Files not moved:')
        print(*files_not_moved, sep='\n')

    print('Complete')
    _ = input('Press Enter to exit\n')


def change_file(path, po_identifier):
    wb = pyxl.load_workbook(filename=path, data_only=True)
    ws = wb.active

    ws['C1'].value = po_identifier
    ws['C1'].alignment = Alignment(wrap_text=False)
    wb.save(Path('export') / (po_identifier + '.xlsx'))


if __name__ == '__main__':
    main()
